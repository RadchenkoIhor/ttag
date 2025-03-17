import pandas as pd # Truth Table Auto Generator (ttag) ver. BetaTest
import openpyxl
import os
import time
import msvcrt
import platform
from colorama import Fore, init

init(autoreset=True)

def live_input():
    vector = ""
    print("\n Введіть функцію, задану вектором значень f: ", end="", flush=True)
    while len(vector) < 16:
        char = msvcrt.getch().decode("utf-8")
        if char.isdigit() and char in "01":
            vector += char
            print(char, end="", flush=True)
            if len(vector) % 4 == 0 and len(vector) < 16:
                print(" ", end="", flush=True)
        elif char == "\x08" and vector:
            if len(vector) % 4 == 0 and len(vector) > 0:
                print("\b \b", end="", flush=True)
            print("\b \b", end="", flush=True)
            vector = vector[:-1]
        elif char == "\r":  # Enter
            break
    print()
    return vector

def get_truth_table(vector: str):
    n = 4
    rows = 2 ** n
    table_data = []
    
    for i in range(rows):
        binary_values = list(format(i, f'0{n}b'))
        table_data.append(binary_values + [vector[i]])
    
    return table_data

def save_to_excel(data, filename: str):
    columns = ['X1', 'X2', 'X3', 'X4', 'F']
    df = pd.DataFrame(data, columns=columns)
    
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        df.to_excel(writer, index=False)
        workbook = writer.book
        worksheet = writer.sheets['Sheet1']
        
        for col in worksheet.iter_cols():
            for cell in col:
                cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
                cell.border = openpyxl.styles.Border(left=openpyxl.styles.Side(style='thin'),
                                                     right=openpyxl.styles.Side(style='thin'),
                                                     top=openpyxl.styles.Side(style='thin'),
                                                     bottom=openpyxl.styles.Side(style='thin'))
        
        header_fill = openpyxl.styles.PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
        for cell in worksheet["A1:E1"][0]:
            cell.fill = header_fill
        
        for row in range(2, worksheet.max_row + 1):
            cell_value = worksheet[f"E{row}"].value
            try:
                int_value = int(cell_value)
                if int_value == 1:
                    worksheet[f"E{row}"].fill = openpyxl.styles.PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
                else:
                    worksheet[f"E{row}"].fill = openpyxl.styles.PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
            except ValueError:
                pass

def open_file(filename: str):
    try:
        os.system(f"start excel {filename}")
        return
    except Exception as e:
        try:
            os.system(f"libreoffice {filename}")
            return
        except Exception as e:
            pass
    
    print(f"Не вдалося відкрити файл. Спробуйте вручну. Помилка: {e}")

def main():
    while True:
        vector = live_input()
        formatted_vector = " ".join([vector[i:i+4] for i in range(0, len(vector), 4)])
        colored_vector = f"{Fore.CYAN}{formatted_vector}{Fore.RESET}"
        confirmation = input(f"\n Ваша функція задана вектором значень: {colored_vector}. Це правильно? ({Fore.LIGHTGREEN_EX}y{Fore.RESET}/{Fore.LIGHTRED_EX}n{Fore.RESET}): ")
        if confirmation.lower() == 'y':
            break
    
    filename = f"f_{formatted_vector.replace(' ', '_')}.xlsx"
    
    colored_filename = f"{Fore.YELLOW}{filename}{Fore.RESET}"
    colored_vector = f"{Fore.CYAN}{formatted_vector}{Fore.RESET}"
    
    print("\n Будую таблицю істинності...")
    table_data = get_truth_table(vector)
    save_to_excel(table_data, filename)
    
    print(f"\n Таблицю істинності для функції f = ({colored_vector}) успішно побудовано та збережено у файл {colored_filename}.")
    
    full_path = os.path.abspath(filename)
    print(f"\n Повний шлях до файлу: {Fore.LIGHTYELLOW_EX}{full_path}{Fore.RESET}.\n")
    
    time.sleep(1)
    open_file(filename)
    input(" Натисніть Enter для виходу...")

if __name__ == "__main__":
    main()
